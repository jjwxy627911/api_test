# -*- coding:utf-8 -*-
# @Time     : 2019-06-25 21:24
# @Author   : davis
# @Email    : jjwxy627911@163.com
# @File     : handle_excel.py
# @Software : PyCharm


from openpyxl import load_workbook
from scripts.constance import CASES_FILE_PATH


class HandleExcel:
    """
    操作excel
    """

    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def get_cases(self):
        """
        获取所有测试用例
        :return:
        """
        self.wb = load_workbook(self.file_path)

        if self.sheet_name is not None:
            sheet = self.wb[self.sheet_name]

        else:

            sheet = self.wb.active

        sheet_head = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        case_rows = list(sheet.iter_rows(min_row=2, values_only=True))

        case_list = []
        for item in case_rows:
            row = dict(zip(sheet_head, item))
            case_list.append(row)

        return case_list

    def write_value(self, row, col, value):
        """
        写入结果
        :param row:
        :param col:
        :param value:
        :return:
        """
        wb = load_workbook(self.file_path)
        if self.sheet_name is not None:
            sheet = wb[self.sheet_name]
        else:
            sheet = wb.active

        sheet.cell(row, col, value)

        wb.save(self.file_path)
        wb.close()

    def close(self):
        self.wb.close()


if __name__ == "__main__":
    xlsx = HandleExcel(CASES_FILE_PATH, "register")
    print(xlsx.get_cases())
    xlsx.close()
    # xlsx.write_value(2, 6, 111)
